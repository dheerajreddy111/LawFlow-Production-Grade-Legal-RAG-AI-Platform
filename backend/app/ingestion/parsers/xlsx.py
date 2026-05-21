"""Excel (.xlsx) parser — sheet- and row-aware, async-safe.

Uses ``openpyxl`` in read-only mode. Multi-sheet workbooks become one
plain-text document with each sheet wrapped in a heading the chunker
recognises as a section break:

    SHEET: <sheet name>
    <header row rendered>
    <each data row rendered>

Per-row rendering matches the CSV parser (``"col: val | col: val"``)
so the same chunker strategy works on both, and a query over tabular
legal data (penalty schedules, jurisdictional bench rosters) gets the
header context with every row.

Sheet and row counts are exposed via :attr:`ParseResult.extra`.
"""

from __future__ import annotations

import asyncio
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.ingestion.parsers.base import BaseParser, ParseResult
from app.integrations.lc import traced


def _stringify(value: Any) -> str:
    """Excel cell → trimmed string. Treats ``None`` and NaN-likes as empty."""
    if value is None:
        return ""
    if isinstance(value, float) and value != value:  # NaN
        return ""
    return str(value).strip()


def _render_row(row: list[Any], header: list[str] | None) -> str:
    cells = [_stringify(v) for v in row]
    if header and len(header) == len(cells):
        return " | ".join(
            f"{(h or '').strip()}: {v}" for h, v in zip(header, cells) if v or (h and h.strip())
        )
    return " | ".join(c for c in cells if c)


def _looks_like_header(row: list[Any]) -> bool:
    if not row:
        return False
    cells = [_stringify(v) for v in row]
    non_empty = [c for c in cells if c]
    if len(non_empty) < max(1, len(cells) // 2):
        return False
    numeric = sum(
        1 for c in non_empty if c.replace(".", "", 1).lstrip("-").isdigit()
    )
    return numeric < len(non_empty) / 2


def _parse_sheet(ws) -> tuple[str, dict[str, Any]]:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    rows = [r for r in rows if any(_stringify(v) for v in r)]
    if not rows:
        return "", {"row_count": 0, "columns": []}

    has_header = _looks_like_header(rows[0])
    header = [_stringify(v) for v in rows[0]] if has_header else None
    data_rows = rows[1:] if has_header else rows

    lines = [_render_row(r, header) for r in data_rows]
    body = "\n\n".join(line for line in lines if line)

    # Add a SHEET heading so the section-aware chunker treats sheets
    # as distinct semantic units in a multi-sheet workbook.
    text = f"SHEET: {ws.title}\n\n{body}" if body else f"SHEET: {ws.title}"
    return text, {
        "row_count": len(data_rows),
        "columns": header or [],
        "has_header": has_header,
        "sheet": ws.title,
    }


class XLSXParser(BaseParser):
    @traced(name="parser.xlsx", run_type="tool")
    async def parse(self, path: Path) -> ParseResult:
        return await asyncio.to_thread(_parse_sync, path)


def _parse_sync(path: Path) -> ParseResult:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        sheet_blocks: list[str] = []
        per_sheet: list[dict[str, Any]] = []
        for ws in wb.worksheets:
            text, meta = _parse_sheet(ws)
            if text.strip():
                sheet_blocks.append(text)
            per_sheet.append(meta)

        full_text = "\n\n".join(sheet_blocks)
        return ParseResult(
            text=full_text,
            page_count=None,
            extra={
                "sheet_count": len(wb.worksheets),
                "sheets": per_sheet,
                "row_count_total": sum(s["row_count"] for s in per_sheet),
            },
        )
    finally:
        wb.close()
